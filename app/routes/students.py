import os
from flask import (Blueprint, render_template, redirect, url_for, flash,
                   request, send_file, current_app)
from flask_login import login_required, current_user
from app.models import Student, ClassSection, FeePayment, Attendance
from app.utils.helpers import generate_reg_no, paginate_query
from app.utils.decorators import staff_required
from app import db
from datetime import date
import io

students_bp = Blueprint('students', __name__, template_folder='../templates')

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@students_bp.route('/')
@login_required
def list_students():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '')
    class_id = request.args.get('class_id', '', type=str)
    status = request.args.get('status', 'active')

    query = Student.query
    if search:
        query = query.filter(
            db.or_(Student.full_name.ilike(f'%{search}%'),
                   Student.reg_no.ilike(f'%{search}%'),
                   Student.father_name.ilike(f'%{search}%'))
        )
    if class_id:
        query = query.filter_by(class_section_id=int(class_id))
    if status:
        query = query.filter_by(status=status)

    pagination = paginate_query(query.order_by(Student.id.desc()), page, 15)
    classes = ClassSection.query.filter_by(is_active=True).all()
    return render_template('students/list.html', students=pagination.items,
                           pagination=pagination, classes=classes,
                           search=search, class_id=class_id, status=status)


@students_bp.route('/add', methods=['GET', 'POST'])
@login_required
@staff_required
def add_student():
    classes = ClassSection.query.filter_by(is_active=True).all()
    if request.method == 'POST':
        student = Student(
            reg_no=generate_reg_no(),
            full_name=request.form.get('full_name'),
            father_name=request.form.get('father_name'),
            mother_name=request.form.get('mother_name'),
            date_of_birth=request.form.get('date_of_birth') or None,
            gender=request.form.get('gender'),
            phone=request.form.get('phone'),
            parent_phone=request.form.get('parent_phone'),
            email=request.form.get('email'),
            address=request.form.get('address'),
            city=request.form.get('city'),
            cnic=request.form.get('cnic'),
            blood_group=request.form.get('blood_group'),
            religion=request.form.get('religion'),
            class_section_id=request.form.get('class_section_id') or None,
            admission_date=request.form.get('admission_date') or date.today(),
        )
        # Handle photo upload
        photo = request.files.get('photo')
        if photo and allowed_file(photo.filename):
            from werkzeug.utils import secure_filename
            filename = f"{student.reg_no}_{secure_filename(photo.filename)}"
            upload_dir = os.path.join(current_app.static_folder, 'uploads', 'students')
            os.makedirs(upload_dir, exist_ok=True)
            photo.save(os.path.join(upload_dir, filename))
            student.photo = f'uploads/students/{filename}'

        db.session.add(student)
        db.session.commit()
        flash(f'Student {student.full_name} added successfully! Reg No: {student.reg_no}', 'success')
        return redirect(url_for('students.view_student', id=student.id))
    return render_template('students/add.html', classes=classes, today=date.today().isoformat())


@students_bp.route('/<int:id>')
@login_required
def view_student(id):
    student = Student.query.get_or_404(id)
    recent_attendance = Attendance.query.filter_by(student_id=id).order_by(
        Attendance.date.desc()).limit(10).all()
    total_att = Attendance.query.filter_by(student_id=id).count()
    present_count = Attendance.query.filter_by(student_id=id, status='present').count()
    att_pct = round((present_count / total_att * 100) if total_att else 0, 1)
    fee_payments = FeePayment.query.filter_by(student_id=id).order_by(
        FeePayment.payment_date.desc()).limit(5).all()
    return render_template('students/view.html', student=student,
                           recent_attendance=recent_attendance,
                           att_pct=att_pct, present_count=present_count,
                           total_att=total_att, fee_payments=fee_payments)


@students_bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@staff_required
def edit_student(id):
    student = Student.query.get_or_404(id)
    classes = ClassSection.query.filter_by(is_active=True).all()
    if request.method == 'POST':
        student.full_name = request.form.get('full_name', student.full_name)
        student.father_name = request.form.get('father_name', student.father_name)
        student.mother_name = request.form.get('mother_name', student.mother_name)
        student.gender = request.form.get('gender', student.gender)
        student.phone = request.form.get('phone', student.phone)
        student.parent_phone = request.form.get('parent_phone', student.parent_phone)
        student.email = request.form.get('email', student.email)
        student.address = request.form.get('address', student.address)
        student.city = request.form.get('city', student.city)
        student.blood_group = request.form.get('blood_group', student.blood_group)
        student.class_section_id = request.form.get('class_section_id') or student.class_section_id
        student.status = request.form.get('status', student.status)

        photo = request.files.get('photo')
        if photo and allowed_file(photo.filename):
            from werkzeug.utils import secure_filename
            filename = f"{student.reg_no}_{secure_filename(photo.filename)}"
            upload_dir = os.path.join(current_app.static_folder, 'uploads', 'students')
            os.makedirs(upload_dir, exist_ok=True)
            photo.save(os.path.join(upload_dir, filename))
            student.photo = f'uploads/students/{filename}'

        db.session.commit()
        flash('Student updated successfully.', 'success')
        return redirect(url_for('students.view_student', id=id))
    return render_template('students/edit.html', student=student, classes=classes)


@students_bp.route('/<int:id>/delete', methods=['POST'])
@login_required
@staff_required
def delete_student(id):
    student = Student.query.get_or_404(id)
    student.status = 'inactive'
    db.session.commit()
    flash(f'Student {student.full_name} has been deactivated.', 'info')
    return redirect(url_for('students.list_students'))


@students_bp.route('/export')
@login_required
@staff_required
def export_students():
    """Export students to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash('openpyxl not installed. Cannot export.', 'danger')
        return redirect(url_for('students.list_students'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students'

    headers = ['Reg No', 'Full Name', "Father's Name", 'Gender', 'Class',
               'Phone', 'Parent Phone', 'Email', 'Address', 'Status', 'Admission Date']
    header_fill = PatternFill(start_color='FF6B35', end_color='FF6B35', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 18

    students = Student.query.all()
    for row, s in enumerate(students, 2):
        class_name = s.class_section.display_name if s.class_section else ''
        ws.append([s.reg_no, s.full_name, s.father_name, s.gender, class_name,
                   s.phone, s.parent_phone, s.email, s.address, s.status,
                   s.admission_date.strftime('%Y-%m-%d') if s.admission_date else ''])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='students_export.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
