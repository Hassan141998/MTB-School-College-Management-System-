from flask import (Blueprint, render_template, request, send_file, flash, redirect, url_for)
from flask_login import login_required
from app.models import (Student, Teacher, Attendance, FeePayment, Mark,
                         Exam, ClassSection, FeeStructure, SalaryRecord)
from app.utils.decorators import staff_required
from app import db
from datetime import date, timedelta
from sqlalchemy import func
import io

reports_bp = Blueprint('reports', __name__, template_folder='../templates')


def _grade_for_pct(pct):
    if pct >= 90: return 'A+'
    if pct >= 80: return 'A'
    if pct >= 70: return 'B'
    if pct >= 60: return 'C'
    if pct >= 50: return 'D'
    return 'F'


@reports_bp.route('/')
@login_required
def index():
    today = date.today()

    # ── Performance: top 10 students by average exam % ──────────────────
    students_active = Student.query.filter_by(status='active').all()
    performance_rows = []
    for s in students_active:
        marks = Mark.query.filter_by(student_id=s.id).all()
        if not marks:
            continue
        total_obtained = sum(m.obtained_marks for m in marks)
        total_possible = sum(m.total_marks for m in marks) or 1
        avg_pct = round(total_obtained / total_possible * 100, 1)
        performance_rows.append({
            'name': s.full_name,
            'class_name': s.class_section.display_name if s.class_section else '—',
            'avg_pct': avg_pct,
            'grade': _grade_for_pct(avg_pct),
        })
    performance_rows.sort(key=lambda r: r['avg_pct'], reverse=True)
    performance_data = performance_rows[:10]

    # ── Financial summary (this month) ───────────────────────────────────
    total_collected = db.session.query(func.sum(FeePayment.total_paid)).scalar() or 0
    tuition_by_class = {
        fs.class_name: fs.amount
        for fs in FeeStructure.query.filter_by(fee_type='tuition', is_active=True).all()
    }
    paid_ids_this_month = db.session.query(FeePayment.student_id).filter(
        FeePayment.month == today.strftime('%B'),
        FeePayment.year == today.year
    ).scalar_subquery()
    defaulters_now = Student.query.filter(
        Student.status == 'active', ~Student.id.in_(paid_ids_this_month)
    ).all()
    total_pending = sum(
        tuition_by_class.get(s.class_section.class_name, 0) if s.class_section else 0
        for s in defaulters_now
    )
    total_students = len(students_active)
    defaulter_count = len(defaulters_now)
    paid_count = total_students - defaulter_count
    financial = {
        'collected': total_collected,
        'pending': total_pending,
        'total_students': total_students,
        'paid_count': paid_count,
        'defaulter_count': defaulter_count,
    }

    # ── Class-wise attendance (this month) ───────────────────────────────
    classes = ClassSection.query.filter_by(is_active=True).all()
    att_data = []
    month_start = today.replace(day=1)
    for cls in classes:
        cls_students = Student.query.filter_by(class_section_id=cls.id, status='active').all()
        if not cls_students:
            continue
        student_ids = [s.id for s in cls_students]
        total_records = Attendance.query.filter(
            Attendance.student_id.in_(student_ids),
            Attendance.date >= month_start, Attendance.date <= today
        ).count()
        present_records = Attendance.query.filter(
            Attendance.student_id.in_(student_ids),
            Attendance.date >= month_start, Attendance.date <= today,
            Attendance.status == 'present'
        ).count()
        avg_pct = round((present_records / total_records * 100) if total_records else 0, 1)
        att_data.append({
            'class_name': cls.display_name,
            'total_students': len(cls_students),
            'present_days': present_records,
            'avg_pct': avg_pct,
        })

    return render_template('reports/index.html', performance_data=performance_data,
                           financial=financial, att_data=att_data)


@reports_bp.route('/students')
@login_required
@staff_required
def student_report():
    classes = ClassSection.query.filter_by(is_active=True).all()
    class_id = request.args.get('class_id')
    students = []
    if class_id:
        students = Student.query.filter_by(
            class_section_id=int(class_id), status='active').order_by(Student.full_name).all()
    student_data = []
    for s in students:
        total_att = Attendance.query.filter_by(student_id=s.id).count()
        present = Attendance.query.filter_by(student_id=s.id, status='present').count()
        att_pct = round((present / total_att * 100) if total_att else 0, 1)
        total_fees = db.session.query(func.sum(FeePayment.total_paid)).filter_by(
            student_id=s.id).scalar() or 0
        student_data.append({
            'student': s, 'att_pct': att_pct, 'total_fees': total_fees,
            'present': present, 'total_att': total_att
        })
    return render_template('reports/performance.html', student_data=student_data,
                           classes=classes, class_id=class_id)


# reports/index.html links to 'reports.student_performance' - register that
# endpoint name for the same view.
reports_bp.add_url_rule('/students', endpoint='student_performance',
                        view_func=student_report)


@reports_bp.route('/financial')
@login_required
@staff_required
def financial_report():
    today = date.today()
    month = int(request.args.get('month', today.month))
    year = int(request.args.get('year', today.year))

    monthly_payments = FeePayment.query.filter(
        func.extract('month', FeePayment.payment_date) == month,
        func.extract('year', FeePayment.payment_date) == year
    ).order_by(FeePayment.payment_date.desc()).all()

    total = sum(p.total_paid for p in monthly_payments)
    by_type = {}
    for p in monthly_payments:
        fee_type = p.fee_structure.fee_type if p.fee_structure else 'other'
        by_type[fee_type] = by_type.get(fee_type, 0) + p.total_paid

    total_salary = db.session.query(func.sum(SalaryRecord.net_salary)).filter(
        SalaryRecord.year == year
    ).scalar() or 0
    net_balance = total - total_salary

    from app.utils.helpers import get_months
    return render_template('reports/financial.html', payments=monthly_payments,
                           total=total, total_collected=total, total_salary=total_salary,
                           net_balance=net_balance, by_type=by_type, month=month, year=year,
                           this_month=today.strftime('%B %Y'),
                           months=get_months(), years=range(2020, today.year + 2))


# reports/index.html links to 'reports.financial' - register that endpoint
# name for the same view.
reports_bp.add_url_rule('/financial', endpoint='financial',
                        view_func=financial_report)


@reports_bp.route('/attendance')
@login_required
@staff_required
def attendance_report():
    classes = ClassSection.query.filter_by(is_active=True).all()
    class_id = request.args.get('class_id')
    from_date = request.args.get('from_date', (date.today() - timedelta(days=30)).isoformat())
    to_date = request.args.get('to_date', date.today().isoformat())

    data = []
    if class_id:
        students = Student.query.filter_by(
            class_section_id=int(class_id), status='active').all()
        for s in students:
            total = Attendance.query.filter(
                Attendance.student_id == s.id,
                Attendance.date >= from_date,
                Attendance.date <= to_date
            ).count()
            present = Attendance.query.filter(
                Attendance.student_id == s.id,
                Attendance.date >= from_date,
                Attendance.date <= to_date,
                Attendance.status == 'present'
            ).count()
            pct = round((present / total * 100) if total else 0, 1)
            data.append({'student': s, 'present': present, 'total': total, 'pct': pct})

    return render_template('reports/attendance.html', data=data, classes=classes,
                           class_id=class_id, from_date=from_date, to_date=to_date)


# Same naming mismatch pattern as the other reports/index.html links.
reports_bp.add_url_rule('/attendance', endpoint='attendance',
                        view_func=attendance_report)


@reports_bp.route('/export/students')
@login_required
@staff_required
def export_student_report():
    """Export student performance report to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        from flask import flash, redirect, url_for
        flash('openpyxl not installed.', 'danger')
        return redirect(url_for('reports.student_report'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Student Performance'
    headers = ['Reg No', 'Name', "Father's Name", 'Class', 'Present', 'Total Days', 'Attendance %']
    hf = PatternFill(start_color='FF6B35', end_color='FF6B35', fill_type='solid')
    hfont = Font(color='FFFFFF', bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hf
        cell.font = hfont
        ws.column_dimensions[cell.column_letter].width = 18

    students = Student.query.filter_by(status='active').all()
    for row, s in enumerate(students, 2):
        total = Attendance.query.filter_by(student_id=s.id).count()
        present = Attendance.query.filter_by(student_id=s.id, status='present').count()
        pct = round((present / total * 100) if total else 0, 1)
        class_name = s.class_section.display_name if s.class_section else ''
        ws.append([s.reg_no, s.full_name, s.father_name, class_name, present, total, pct])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='student_performance.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# Same naming mismatch pattern as the other reports/index.html links.
reports_bp.add_url_rule('/export/performance', endpoint='export_performance',
                        view_func=export_student_report)


@reports_bp.route('/export/attendance')
@login_required
@staff_required
def export_attendance():
    """Export class-wise attendance summary (this month) to Excel.
    Covers both reports/index.html's 'export_attendance' link and
    reports/attendance.html's 'attendance_excel' link (same report)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash('openpyxl not installed.', 'danger')
        return redirect(url_for('reports.attendance_report'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Attendance Summary'
    headers = ['Class', 'Total Students', 'Present Days', 'Attendance %']
    hf = PatternFill(start_color='FF6B35', end_color='FF6B35', fill_type='solid')
    hfont = Font(color='FFFFFF', bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hf
        cell.font = hfont
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 20

    today = date.today()
    month_start = today.replace(day=1)
    classes = ClassSection.query.filter_by(is_active=True).all()
    for row, cls in enumerate(classes, 2):
        cls_students = Student.query.filter_by(class_section_id=cls.id, status='active').all()
        if not cls_students:
            continue
        student_ids = [s.id for s in cls_students]
        total_records = Attendance.query.filter(
            Attendance.student_id.in_(student_ids),
            Attendance.date >= month_start, Attendance.date <= today
        ).count()
        present_records = Attendance.query.filter(
            Attendance.student_id.in_(student_ids),
            Attendance.date >= month_start, Attendance.date <= today,
            Attendance.status == 'present'
        ).count()
        pct = round((present_records / total_records * 100) if total_records else 0, 1)
        ws.append([cls.display_name, len(cls_students), present_records, pct])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='attendance_summary.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# reports/attendance.html links to 'reports.attendance_excel' for the same export.
reports_bp.add_url_rule('/export/attendance-excel', endpoint='attendance_excel',
                        view_func=export_attendance)

# reports/performance.html links to 'reports.performance_excel' for the
# same student-performance export.
reports_bp.add_url_rule('/export/performance-excel', endpoint='performance_excel',
                        view_func=export_student_report)
